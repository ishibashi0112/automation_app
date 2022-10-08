import base64
from dataclasses import dataclass
import io
import re
from typing import TypedDict
from pdfminer.high_level import extract_text
from openpyxl import load_workbook
from server.classes.StmpMailer import StmpMailer
from server.function import error_action, get_list_value, success_action, today, today_str
from PyPDF2 import PdfFileWriter, PdfFileReader
from pathlib import Path
from server.type import MainProcessingResultsType

class CMPOExcelData(TypedDict):
    w_c_code: str
    w_c_name: str
    key_person: str
    email: str
    sub_person: str
    sub_email: str


@dataclass
class CreateMailForPaperOrder:
    pdfFile: str

    NEW_FILE_BASE_PATH = "Y:\\530_資材事業課\\パーツセンター\\※GPC_購買部\\発注G\\automation\\paper_order_sheet"
    MAIN_DRAWINGS_FOLDER_PATH = "W:\\01_図面"
    OLD_DRAWINGS_FOLDER_PATH = "W:\\02_文書検索\\4 図面\\201 図面\\01資材確認\\01必要"
    TOSHIBA_DRAWINGS_FOLDER_PATH = "W:\\02_文書検索\\4 図面\\202 図面≪東芝≫"

    def __post_init__(self) -> None:
        self.wb = load_workbook("未EDI企業一覧 (1).xlsx")
        self.ws = self.wb.active
        self.last_rows = self.get_last_rows_index()   

        decoded_pdf_files = base64.b64decode(self.pdfFile)
        self.bytesIO_pdf_files = io.BytesIO(decoded_pdf_files)

    def get_last_rows_index(self) -> int:
        current_rows = 2
        while True:
            current_cell_value = self.ws.cell(current_rows, 1).value
            if not current_cell_value:
                break

            current_rows += 1
        
        return current_rows -1


    def get_required_excel_data(self, rows: int) ->  CMPOExcelData:
        required_excel_data: CMPOExcelData = {
            "w_c_code": self.ws.cell(rows, 1).value,
            "w_c_name": self.ws.cell(rows, 2).value,
            "key_person": self.ws.cell(rows, 4).value,
            "email": self.ws.cell(rows, 5).value,
            "sub_person": self.ws.cell(rows, 6).value,
            "sub_email": self.ws.cell(rows, 7).value,
        }

        return required_excel_data

    def send_order_mail(self, excel_data: CMPOExcelData, attach_files: list[str]) -> None:
        from_address = "komori.service.automation@gmail.com"
        password = "vwsogyjfmxjoqzmf"
        to_address = 'yuki_ishibashi@komori.co.jp'

        subject = f"{today_str()} 新規特急製作依頼書の送付"
        body = f"""
        {excel_data['email'] if excel_data["email"] else ""}
        {excel_data['sub_email'] if excel_data["sub_email"] else ""}

        {excel_data["w_c_name"]}
        {excel_data["key_person"]}様　{f"{excel_data['sub_person']}様" if excel_data["sub_person"] else ""}

        いつも大変お世話になっております。
        本件、添付にて新規の特急製作依頼書を添付いたしました。

        ご確認の程宜しく御願いいたします。

        小森ｺｰﾎﾟﾚｰｼｮﾝ　つくばｸﾞﾛｰﾊﾞﾙﾊﾟｰﾂｾﾝﾀｰ　石橋
        """
        
        
        mail = StmpMailer("smtp.gmail.com", 587, from_address, password)
        mail.set_send_to(from_address, to_address, subject=subject )
        mail.set_body(body)
        
        for file_path in attach_files:
            mail.attach_file(file_path, Path(file_path).name)

        mail.send_email()


    def create_clone_pdf_page(self, new_file_path: str, page_idx: int) -> None:
        org_pdf = PdfFileReader(self.bytesIO_pdf_files)
        org_pdf.numPages
        new_pdf = PdfFileWriter()
        new_pdf.addPage(org_pdf.getPage(page_idx))

        with open(new_file_path, "wb") as f:
            new_pdf.write(f)





def create_mail_for_paper_order(pdfFile: str) -> MainProcessingResultsType:
    try:
        cmpo = CreateMailForPaperOrder(pdfFile)

        for i in range(100):
            pdf_text = extract_text(cmpo.bytesIO_pdf_files, page_numbers=[i])

            if not pdf_text:
                break

            removed_blanks_text = pdf_text.replace(" ", "")
            pdf_text_list = removed_blanks_text.split()

            work_code_pattern = r"([0-9]{5})"
            filter_supplier_code = list(filter(lambda text: re.fullmatch(work_code_pattern, text), pdf_text_list))
            supplier_code = get_list_value(filter_supplier_code, 0)

            item_num_pattern = r"([A-Z0-9]{10})"
            item_num_list = list(filter(lambda text: re.fullmatch(item_num_pattern, text), pdf_text_list))

            new_file_path = f"{CreateMailForPaperOrder.NEW_FILE_BASE_PATH}\\特急製作依頼書_{supplier_code}_{today().strftime('%Y年%m月%d日%H時%M分%S秒')}.pdf"
            cmpo.create_clone_pdf_page(new_file_path, i)

            # 添付図面を取得
            attach_files = [new_file_path]
            for item_num in item_num_list:

                not_pattrens = [ re.compile(pattren) for pattren in [r"^3Z", r"^5[A-Z]"]]
                if not any([pattren.search(item_num) for pattren in not_pattrens]):

                    main_file_path = f"{CreateMailForPaperOrder.MAIN_DRAWINGS_FOLDER_PATH}\\{item_num}.pdf"
                    old_file_path = f"{CreateMailForPaperOrder.OLD_DRAWINGS_FOLDER_PATH}\\図面{item_num}.pdf"
                    toshiba_file_path = f"{CreateMailForPaperOrder.TOSHIBA_DRAWINGS_FOLDER_PATH}\\図面東芝{item_num[:6]}.pdf"

                    if Path(main_file_path).exists():
                        attach_files.append(main_file_path)

                    elif Path(old_file_path).exists():
                        attach_files.append(old_file_path)
                        
                    elif Path(toshiba_file_path).exists():
                        attach_files.append(toshiba_file_path)
                        
                
                    
                    

            for i in range(cmpo.last_rows): 
                excel_data = cmpo.get_required_excel_data(i+2)
                excel_supplier_code = excel_data["w_c_code"]

                if not excel_supplier_code:
                    break

                if supplier_code == excel_supplier_code:
                    cmpo.send_order_mail(excel_data, attach_files)

                    break
        
        for file in Path(CreateMailForPaperOrder.NEW_FILE_BASE_PATH).iterdir():
            file.unlink()

        return success_action()
    except Exception as e:
        return error_action(e)


if __name__ == "__main__":
    print(123)




# def split_pdf(src_path, new_basepath):
#     org_pdf = PdfFileReader(src_path)
#     for i in range(org_pdf.numPages):
#         new_pdf = PdfFileWriter()
#         new_pdf.addPage(org_pdf.getPage(i))
#         i = "{0:02}".format(i)
#         with open('{}_{}.pdf'.format(new_basepath, i), 'wb') as f:
#             new_pdf.write(f)