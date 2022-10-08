from dataclasses import dataclass
import io
import os
import signal
import base64

from dotenv import load_dotenv
from server.classes.Op import Op
from server.classes.OpKoutei import OpKoutei
from server.classes.OpPayments import OpPayments
from server.type import MainProcessingResultsType, OpExcelRequiredDataType
from server.function import error_action, success_action
from openpyxl import load_workbook
from server.utils import get_id

# load_dotenv("../../.env")

    
@dataclass
class OpExcel(Op):
    id: str
    password: str
    excel: str 

    def __post_init__(self):
        decoded_excel_file = base64.b64decode(self.excel)
        BytesIO_excel_file = io.BytesIO(decoded_excel_file)
        wb = load_workbook(BytesIO_excel_file)
        self.ws = wb.active
        super().set_up()
        super().open_main_page(os.environ["SERVICE_SYSTEM_URL"])
        super().Login(self.id, self.password)
        
    
    def header_input(self, order_num: str, item_num: str ) -> None:
        super().set_value(get_id("受注NO_op_search"), order_num)
        super().set_value(get_id("品番_op_search"), item_num)
        super().btn_click(get_id("検索_op_search"))
        
        super().nomal_wait()

    
    def get_last_rows_index(self) -> int:
        current_rows = 2
        while True:
            current_cell_value = self.ws.cell(current_rows, 1).value
            if not current_cell_value:
                break

            current_rows += 1
        
        return current_rows
    
    def get_required_excel_data(self, rows: int) -> OpExcelRequiredDataType:
        way = self.ws.cell(rows, 1).value
        order_num = self.ws.cell(rows, 3).value
        lines_num = self.ws.cell(rows, 4).value
        item_num = self.ws.cell(rows, 6).value
        qty = self.ws.cell(rows, 9).value
        delivery_time = self.ws.cell(rows, 17).value
        comment = self.ws.cell(rows, 24).value

        return {
            "対応": way,
            "受注NO": order_num,
            "行NO": lines_num,
            "品番": item_num,
            "数量": qty,
            "納期": delivery_time,
            "備考": comment
        }
    
    def complete_process(self, current_excel_data: OpExcelRequiredDataType) -> bool:
        items_list = super().get_item_els()
        is_order_exist = False
        for i, item in enumerate(items_list):
            #次にみるitemが無い場合
            if item == []:
                break

            if super().get_value(get_id("受注NO_op_results", i)) == current_excel_data["受注NO"] and super().get_value(get_id("行NO_op_results", i)) == current_excel_data["行NO"]:

                super().btn_click(get_id("更新_op_results", i))
                super().set_value(get_id("備考_op_results", i), current_excel_data["備考"])
                super().set_qty(i, current_excel_data["数量"])
                

                if current_excel_data["対応"] == "削除":
                    super().btn_click(get_id("削除_op_results", i))


                elif current_excel_data["対応"] == "内示":
                    if not ("紙" in current_excel_data["備考"] or "支給" in current_excel_data["備考"]):
                        super().btn_click(get_id("内示_op_results", i))
                    
                    if super().get_value(get_id("製番_op_results", i))[:2] == "93":
                        super().set_value(get_id("製番_op_results", i), "9300NAI")

                    koutei = OpKoutei(self.brower, self.wait, i, delivery_time=current_excel_data["納期"])
                    koutei.from_excel_to_process()

                    payments = OpPayments(self.brower, self.wait, i)
                    payments.click_only()
                    
                else: 
                    super().btn_click(get_id("確定check_op_results", i))
                    
                    koutei = OpKoutei(self.brower, self.wait, i, delivery_time=current_excel_data["納期"])
                    koutei.click_only()

                    payments = OpPayments(self.brower, self.wait, i)
                    payments.click_only()
                

                is_result = super().complete()
                if not is_result:
                    break

                is_order_exist = True
                break

        return is_order_exist




def op_excel(id: str, password: str, excel: str) -> MainProcessingResultsType:
    op = OpExcel(id, password, excel)
    try:
        op.Login(id, password)
        op.menu_open("op_entry", 1)
        op.screen_switching(1)
        
        for i in range(op.get_last_rows_index()):
            current_excel_data = op.get_required_excel_data(i+2)
            
            if not current_excel_data["受注NO"]:
                continue

            op.header_input(current_excel_data["受注NO"], current_excel_data["品番"])
            
            if not len(op.get_elements("id", "M1")):
                continue

            if op.check_pages_exists():
                
                is_not_last_page = True
                is_order_exist = False
                while is_not_last_page:

                    is_order_exist = op.complete_process(current_excel_data)
                    if is_order_exist:
                        break

                    is_not_last_page = op.check_next_page_exists()
                    if is_not_last_page:
                        op.move_next_page()               
                        op.nomal_wait()
                
                if not is_order_exist:
                   op.clear_search_input()

                
            else:
                is_order_exist = op.complete_process(current_excel_data)

                if not is_order_exist:       
                    op.clear_search_input()
                  
                
                
                
# ========================================================================================
        return success_action()
    except Exception as e:
        return error_action(e)
    finally:
        os.kill(op.brower.service.process.pid, signal.SIGTERM)



    
if __name__ == "__main__":
    op_excel("kkc4726", "kkc@4726", "")