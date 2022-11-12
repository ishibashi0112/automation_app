# C:\Users\kkc4726\Desktop\projects\automation\repair_order

from dataclasses import dataclass, field
import os
import signal
from typing import Any, Final, Literal, TypedDict
from dotenv import load_dotenv
from openpyxl import load_workbook
from server.classes.Obi import Obi
from server.classes.Op import Op
from server.classes.OpKoutei import OpKoutei
from server.classes.OpPayments import OpPayments
from server.function import datetime_to_str, error_action, get_various_weeks, list_to_merge_str, success_action, today_str
from server.type import ExcelDataObi, KouteiDataType, MainProcessingResultsType, RepairOrderItemsType
from server.utils import get_id



class RepairOrderExcelListType(TypedDict):
    order_num: str
    lines_num: str
    item_num: str
    rows: int
    obi_values: ExcelDataObi
    KOUTEI_data: list[KouteiDataType]



@dataclass
class RepairOrder(Op):

    id: str
    password: str
    orders: list[RepairOrderItemsType]
    excel_data_list: list[RepairOrderExcelListType] = field(default_factory=list)

    file_name = os.environ["REPAIR_ORDER_EXCEL_FILE_PATH"]

    def __post_init__(self) -> None:
        super().set_up()
        super().open_main_page(os.environ["SERVICE_SYSTEM_URL"])
        super().Login(self.id, self.password)

    def open_excel(self) -> bool:
        wb = load_workbook(RepairOrder.file_name)
        ws = wb["預かり修理品 2018.4.20～"]
        self.wb = wb
        self.ws = ws
        try:
            wb.save(RepairOrder.file_name)
            return True
        except PermissionError:
            return False
            
    
    def excel_list_append(self, order_num: str, lines_num: str, item_num: str, rows: int, excel_values: ExcelDataObi, KOUTEI_data: list[KouteiDataType]) -> None:
        excel_lib: RepairOrderExcelListType = {
            "order_num": order_num,
            "lines_num": lines_num,
            "item_num": item_num,
            "rows": rows,
            "obi_values": excel_values,
            "KOUTEI_data": KOUTEI_data,
        }

        self.excel_data_list.append(excel_lib)


    def write_excel(self) -> None:
        value_keys: Final[list[str]] = ["受注担当者","品名","仕様","受注数","客先code","客先名","製番",]
        cell_idx: Final[list[int]]  = [5,7,8,9,10,11,12]

        for values in self.excel_data_list:
            rows = values["rows"]
            KOUTEI_data = values["KOUTEI_data"]
            obi_values = values["obi_values"]
            self.ws.cell(rows, 2).value = values["order_num"] 
            self.ws.cell(rows, 3).value = values["lines_num"] 
            self.ws.cell(rows, 6).value = values["item_num"] 
            self.ws.cell(rows, 13).value = list_to_merge_str(KOUTEI_data, "/", "取引先code")
            self.ws.cell(rows, 14).value = list_to_merge_str(KOUTEI_data, "/", "取引先名")
            self.ws.cell(rows, 15).value = KOUTEI_data[-1]["納期"]

            for key, idx in zip(value_keys, cell_idx):
                self.ws.cell(rows, idx).value = obi_values.get(key)
        

        self.wb.save(RepairOrder.file_name)
    
    def get_start_rows(self) -> int:
        rows = 1
        process_start_rows = True
        while process_start_rows:
            current_cell = self.ws.cell(rows, 2)
            current_value = current_cell.value
            if not current_value:
                process_start_rows = False
                break

            rows = rows + 1
        return rows
    
    def op_search(self, order_num: str, item_num: str) -> bool:
        super().set_value(get_id("受注NO_op_search"), order_num)
        super().set_value(get_id("品番_op_search"), item_num)
        super().btn_click(get_id("検索_op_search"))
        super().nomal_wait()
        search_result = len(super().get_elements("id", "M1"))

        return True if search_result else False
    
    def KOUTEI_process(self, i: int) -> list[KouteiDataType]:
        koutei = OpKoutei(self.brower, self.wait, i)
        koutei.start()
        koutei.set_price_not_setting()

        weeks: Literal[4, 3, 2] =  4 if koutei.num == 1 else 3 if koutei.num == 2 else 2
        delivery_time = get_various_weeks("move", weeks)
        # delivery_time_str = datetime_to_str(delivery_time)                

        for i in range(koutei.num):
            delivery_time_str = datetime_to_str(delivery_time)    
            super().set_value(get_id("納期_op_KOUTEI", i), delivery_time_str)
            if koutei.num > 1:
                move_weeks_num = 3 if koutei.num == 2 else 2 
                delivery_time = get_various_weeks("move", move_weeks_num, delivery_time)

        KOUTEI_data = koutei.get_data()

        koutei.end()

        return KOUTEI_data
    
    def payments_process(self, i):
        payments = OpPayments(self.brower, self.wait, i)
        payments.click_only()


def repair_order(id: str, password: str, orders: list[RepairOrderItemsType]) -> MainProcessingResultsType:

    ro = RepairOrder(id, password, orders)
    try:
        opened_file = ro.open_excel()
        if not opened_file:
          raise PermissionError("誰かが既に開いている可能性があります")

        ro.menu_open("op_entry", 1)
        ro.menu_open("obi_entry", 2)
        ro.screen_switching(1)

        current_rows = ro.get_start_rows()
        for order in ro.orders:
            excel_order_num = order["orderNum"]
            excel_lines_num = order["linesNum"]
            excel_item_num = order["itemNum"]

            if (not excel_order_num) or (not excel_lines_num) or (not excel_item_num):
                continue
            
            search_result = ro.op_search(excel_order_num, excel_item_num)

            if not search_result:
                current_rows += 1
                continue
            

            items_list = ro.get_item_els()
            for i, item in enumerate(items_list):
                #次にみるitemが無い場合
                if not len(item):
                    ro.clear_search_input()
                    break

                # 受注NOと行NOが一致しない場合
                current_op_order_num = ro.get_value(get_id("受注NO_op_results", i))
                current_op_lines_num = ro.get_value(get_id("行NO_op_results", i))
                if not (current_op_order_num == excel_order_num and current_op_lines_num == excel_lines_num):
                    continue

                obi = Obi(ro.brower, ro.wait, excel_order_num, excel_lines_num, excel_item_num, 2)
                obi_search_result = obi.search()

                if not obi_search_result:
                    ro.screen_switching(1)
                    ro.clear_search_input()
                    break

                result = obi.check_all_order()
                obi_result = result["data"]
                obi_excel_data = result["excel"]
                
                
                ro.screen_switching(1)

                if not obi_result["exist"]:
                    ro.delete(i, "受注残無し")
                    break        
            
                ro.btn_click(get_id("更新_op_results", i))
                ro.set_value(get_id("仕様_op_results", i), obi_excel_data["客先名"])
                ro.btn_click(get_id("内示_op_results", i))


                ro.payments_process(i)
                
                KOUTEI_data = ro.KOUTEI_process(i)
                KOUTEI_num = len(KOUTEI_data)
    
               
                if KOUTEI_num > 1:
                    KOUTEI_str_list = [f"工程({i+1}):{data['取引先名']}様" for i, data in enumerate(KOUTEI_data)]
                    comment = f"{today_str()} {' '.join(KOUTEI_str_list)}  預かり修理" 
                    ro.set_value(get_id("備考_op_results", i), comment)
                else:
                    ro.set_value(get_id("備考_op_results", i), f"{today_str()} 預かり修理")

    
                # ro.write_excel(excel_order_num, excel_lines_num, excel_item_num, current_rows, obi_excel_data, KOUTEI_data)
                
                ro.excel_list_append(excel_order_num, excel_lines_num, excel_item_num, current_rows, obi_excel_data, KOUTEI_data)

                ro.complete()

                break

            current_rows = current_rows + 1
        
        ro.write_excel()

        return success_action()

    except PermissionError  as e:
        return error_action(e)

    except Exception as e:
        ro.write_excel()
        return error_action(e)
        
    finally:
        os.kill(ro.brower.service.process.pid, signal.SIGTERM)
        

         
if __name__ == "__main__":
    print("repair_order")
    # repair_order("kkc4726","kkc@4726")

        



   