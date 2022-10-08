# C:\Users\kkc4726\Desktop\projects\automation\repair_order

from dataclasses import dataclass
import io
import base64
import os
import re
from glob import glob
from pdfminer.high_level import extract_text
from reportlab.pdfgen import canvas
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import A4, landscape
from PyPDF2 import PdfFileWriter, PdfFileReader
from openpyxl import load_workbook
from server.function import error_action, get_list_value, success_action, today_str
from server.type import InsertExcelRequiredDataType




@dataclass
class insertOrderInfo:
    pdfFiles: str
    excelFile: str

    def __post_init__(self):
        decoded_pdf_files: bytes = base64.b64decode(self.pdfFiles)
        self.pdf_files: io.BytesIO = io.BytesIO(decoded_pdf_files)
        decoded_excel_file: bytes = base64.b64decode(self.excelFile)
        BytesIO_excel_file = io.BytesIO(decoded_excel_file)
        wb = load_workbook(BytesIO_excel_file)
        self.ws = wb.active
        self.last_rows = self._get_last_rows_index()
    
    def _get_last_rows_index(self) -> int:
        current_rows = 2
        while True:
            current_cell_value = self.ws.cell(current_rows, 1).value
            if not current_cell_value:
                break

            current_rows += 1
        
        return current_rows -1

    def get_required_excel_data(self, rows: int) -> InsertExcelRequiredDataType:
        required_excel_data = {
            "取引先code": self.ws.cell(rows, 1).value,
            "品番": self.ws.cell(rows, 2).value,
            "受注NO": self.ws.cell(rows, 3).value,
            "行NO": self.ws.cell(rows, 4).value,
            "納期": self.ws.cell(rows, 5).value,
        }

        return required_excel_data
    


def get_latest_modified_file_path(dirname):
    target = os.path.join(dirname, '*')
    files = [(f, os.path.getmtime(f)) for f in glob(target)]
    latest_modified_file_path = sorted(files, key=lambda files: files[1])[-1]
    return latest_modified_file_path[0]



def insert_order_info(pdfFiles: str, excelFile: str ):
    try:
        insert = insertOrderInfo(pdfFiles, excelFile)

        pdf_to_text_pages_list = []
        for i in range(100):
            pdf_text = extract_text(insert.pdf_files, page_numbers=[i])
            if not pdf_text:
                break
            # 抽出したテキストの空白を全て削除して、改行区切りでリスト化する
            remove_blanks_text = pdf_text.replace(" ", "")
            pdf_text_list = remove_blanks_text.split()

            print("=======================================================================")

            print("------first_pdf_txt---------- ")
            print(pdf_text_list)
            print("----------------------------- ")
            
            work_code_pattern = r"([0-9]{5})"
            filter_supplier_code = list(filter(lambda text: re.fullmatch(work_code_pattern, text), pdf_text_list))
            supplier_code = get_list_value(filter_supplier_code, 0)

            print("------supplier_code---------- ")
            print(supplier_code)
            print("----------------------------- ")
            
            order_info_pattern = r"(0000)([4-9]{1})([0-9]{5})"
            filter_order_info = filter(lambda text: re.match(order_info_pattern, text), pdf_text_list)
            order_info_list = list(
                map(lambda order_info: {"受注NO": order_info[:10], "行NO": order_info[10:]},
                    filter_order_info
                ))
            
            print("------order_info_list---------- ")
            print(order_info_list)
            print("----------------------------- ")

            item_num_pattern = r"([A-Z0-9]{10})"
            filter_item_num = list(filter(lambda text: re.fullmatch(item_num_pattern, text), pdf_text_list))
            item_num_list = []
            if not len(filter_item_num) == len(order_info_list):
                def check_item_num_exists_in_excel(item_num):
                    for i in range(insert.last_rows): 
                        excel_data = insert.get_required_excel_data(i+2)

                        if not excel_data["受注NO"]:
                            break

                        excel_supplier_code = excel_data["取引先code"].split("__")
                        if item_num == excel_data["品番"] and supplier_code in excel_supplier_code:
                        # supplier_code == excel_data["取引先code"]:
                            return True
                    return False

                item_num_list = list(filter(check_item_num_exists_in_excel, filter_item_num))
            else:
                item_num_list = filter_item_num
            
            print("------item_num_list---------- ")
            print(item_num_list)
            print("----------------------------- ")
               
            print("=======================================================================")
                    
            delivery_time_list = []
            for item_num, order_info in zip(item_num_list, order_info_list):
                for i in range(insert.last_rows): 
                    excel_data = insert.get_required_excel_data(i+2)

                    if not excel_data["受注NO"]:
                        break
                    
                    excel_supplier_code_list = excel_data["取引先code"].split("__")
                    delivery_time_str_list = excel_data["納期"].split("__")

                    if order_info["行NO"]:
                        if (supplier_code in excel_supplier_code_list
                            and excel_data["受注NO"] == order_info["受注NO"] 
                            and excel_data["行NO"] == order_info["行NO"] 
                            and excel_data["品番"] == item_num):

                            supplier_code_list_idx = excel_supplier_code_list.index(supplier_code)
                            delivery_time_str = delivery_time_str_list[supplier_code_list_idx]
                            sliced_delivery_time = delivery_time_str[5:]

                            print(f"excel_supplier_code_list:  {excel_supplier_code_list}")
                            print(f"delivery_time:  {sliced_delivery_time}")

                            delivery_time_list.append(sliced_delivery_time)
                    
                    else:
                        if (supplier_code in excel_supplier_code_list
                            and excel_data["受注NO"] == order_info["受注NO"] 
                            and excel_data["品番"] == item_num):

                            supplier_code_list_idx = excel_supplier_code_list.index(supplier_code)
                            delivery_time_str = delivery_time_str_list[supplier_code_list_idx]
                            sliced_delivery_time = delivery_time_str[5:]

                            print(f"excel_supplier_code_list:  {excel_supplier_code_list}")
                            print(f"delivery_time:  {sliced_delivery_time}")

                            delivery_time_list.append(sliced_delivery_time)

                            break
                
            pdf_to_text_pages_list.append(delivery_time_list)

        DOWNLOAD_PATH = f"{os.getenv('HOMEDRIVE')}{os.getenv('HOMEPATH')}\\Downloads"
        insert_text_file_path = f"{DOWNLOAD_PATH}\\{today_str(True)}_A.pdf"
        # PDF新規作成
        cc = canvas.Canvas(insert_text_file_path, pagesize=landscape(A4))   
        # フォントの設定
        font_name = "HeiseiKakuGo-W5"
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))
        cc.setFont(font_name ,size=12)
        # 挿入位置
        target_x, target_y = 515, 355

        print(f"pdf_to_text_pages_list :{pdf_to_text_pages_list}")       
        for delivery_time_list in pdf_to_text_pages_list:
            for delivery_time in delivery_time_list:
               # 文字列挿入
                
                cc.drawString(target_x, target_y, delivery_time)
                target_y -= 24 
            # ページ挿入
            target_y = 355
            cc.showPage()
        # 保存
        cc.save()

        text_file = PdfFileReader(insert_text_file_path, "rb")
        input_file = PdfFileReader(insert.pdf_files, strict=False)
        page_count = input_file.getNumPages()

        # 新規の出力ファイル作成
        output_file = PdfFileWriter()
        # 既存の全体ページをループで回す
        for page_num in range(page_count):

            input_page = input_file.getPage(page_num)
            # 既存のページとテキストをmergeする
            input_page.mergePage(text_file.getPage(page_num))
            # 圧縮する
            input_page.compressContentStreams()
            # 出力ファイルにページを追加する
            output_file.addPage(input_page)
        # 出力ファイル保存
        with open(f"{insert_text_file_path[:-6]}_new.pdf", "wb") as outputStream:
            output_file.write(outputStream)
        

        return success_action()
    except Exception as e:
        return error_action(e)
   

  
     
if __name__ == "__main__":
    print(123)
   