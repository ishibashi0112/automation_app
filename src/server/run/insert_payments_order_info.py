# C:\Users\kkc4726\Desktop\projects\automation\repair_order

from dataclasses import dataclass
from functools import reduce
import io
import base64
from typing import TypedDict
import os
from re import fullmatch
from pdfminer.high_level import extract_text
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfbase.ttfonts import TTFont
from PyPDF2 import PdfFileWriter, PdfFileReader
from openpyxl import load_workbook
from server.function import error_action, success_action, today_str

class IPOExcelRequiredDataType(TypedDict):
    受注NO: str
    行NO: str
    親品番: str
    支給品番: str


class orderNumInfo(TypedDict):
    order_num: str
    idx: int



@dataclass
class insertPaymentsOrderInfo:
    pdfFiles: str
    excelFile: str

    def __post_init__(self):
        decoded_pdf_files: bytes = base64.b64decode(self.pdfFiles)
        self.pdf_files: io.BytesIO = io.BytesIO(decoded_pdf_files)
        decoded_excel_file: bytes = base64.b64decode(self.excelFile)
        BytesIO_excel_file = io.BytesIO(decoded_excel_file)
        wb = load_workbook(BytesIO_excel_file)
        self.ws = wb.active
        self.last_rows = self._get_last_rows_index()
    
    def _get_last_rows_index(self) -> int:
        current_rows = 2
        while True:
            current_cell_value = self.ws.cell(current_rows, 1).value
            if not current_cell_value:
                break

            current_rows += 1
        
        return current_rows

    def get_required_excel_data(self, rows: int) -> IPOExcelRequiredDataType:
        required_excel_data: IPOExcelRequiredDataType = {
            "受注NO": self.ws.cell(rows, 1).value,
            "行NO": self.ws.cell(rows, 2).value,
            "親品番": self.ws.cell(rows, 3).value,
            "支給品番": self.ws.cell(rows, 10).value,
        }

        return required_excel_data



def insert_payments_order_info(pdfFiles: str, excelFile: str ):
    try:

        DOWNLOAD_PATH = f"{os.getenv('HOMEDRIVE')}{os.getenv('HOMEPATH')}\\Downloads"
        insert_text_file_path = f"{DOWNLOAD_PATH}\\支給品{today_str(True)}_A.pdf"
        # PDF新規作成
        cc = canvas.Canvas(insert_text_file_path, pagesize=portrait(A4))   
        # # フォントの設定
        # font_name = "BIZ-UDGothicR"
        # pdfmetrics.registerFont(TTFont("BIZ-UDGothicR", r"C:\Windows\Fonts\BIZ-UDGothicR.ttc"))
        # cc.setFont(font_name ,size=10)
        # 挿入位置
        target_x, target_y = 500, 635


        ipo = insertPaymentsOrderInfo(pdfFiles, excelFile)

        for i in range(100):
            pdf_text = extract_text(ipo.pdf_files, page_numbers=[i])

            if not pdf_text:
                break

            remove_blanks_text = pdf_text.replace(" ", "")
            pdf_text_list: list[str] = remove_blanks_text.split()
            
            order_num_pattern = r"(0000)([4-9]{1})([0-9]{5})"
            item_num_pattern = r"([A-Z0-9]{10})"
            only_uppercase_letter_text_pattern = r"([A-Z]{10})"

            parent_item_num = pdf_text_list[2]
            order_num_info: list[orderNumInfo] = [{"order_num": text, "idx": i} for i, text in enumerate(pdf_text_list) if fullmatch(order_num_pattern, text) ]
            order_num_idx = order_num_info[0]["idx"]
            order_num = order_num_info[0]["order_num"]

            lines_num =  pdf_text_list[order_num_idx + 1]
            

            remove_parent_item_num = pdf_text_list[4:]
            remove_non_item_num_text  = list(filter(lambda text: fullmatch(item_num_pattern, text), remove_parent_item_num))
            remove_order_num  = [text for text in remove_non_item_num_text if not fullmatch(order_num_pattern, text)]
            remove_only_uppercase_letter_text: list[str]  = [text for text in remove_order_num if not fullmatch(only_uppercase_letter_text_pattern, text)]
            payments_item_list: list[str] = reduce(
                lambda prev, current: [*prev]   if current in prev else [*prev, current] 
                , remove_only_uppercase_letter_text
                , []
            )


            print("==========================================================================================")
            print(f"{i+1}ﾍﾟｰｼﾞ目")
            print("  ")
            print(pdf_text_list)
            print("  ")
            print(f"order_num: {order_num}")
            print(f"lines_num: {lines_num}")
            print(f"parent_item_num {parent_item_num}")
            print(f"payments_item {payments_item_list}")
            print("==========================================================================================")


            for row in range(ipo.last_rows):
                excel_data = ipo.get_required_excel_data(row + 2)

                if excel_data["親品番"] == parent_item_num:
                    parent_supplier_name = ipo.ws.cell(row + 2, 7).value
                          
                    font_name = "BIZ-UDGothicR"
                    pdfmetrics.registerFont(TTFont("BIZ-UDGothicR", r"C:\Windows\Fonts\BIZ-UDGothicR.ttc"))
                    cc.setFont(font_name ,size=10)
                    cc.drawString(120, 690, parent_supplier_name)
                    break





            for item_num in payments_item_list:
                print("==========================================================================================")
                print(f"payments_item: {item_num}")
              
                
                # excel繰り返し
                for row in range(ipo.last_rows):
                    excel_data = ipo.get_required_excel_data(row + 2)

                    if not excel_data["受注NO"]:
                        break
                    
                    if (excel_data["受注NO"] == order_num 
                        and excel_data["行NO"] == lines_num 
                        and excel_data["親品番"] == parent_item_num
                        and excel_data["支給品番"] == item_num ):
                        print("OK")
                  
                       
                        supplier_name = ipo.ws.cell(row + 2, 12).value
                        # 文字列挿入
                        if supplier_name:
                            replace_key = ["㈱", "(株)","（株）", "㈲", "(有)","（有）", "株式会社", "有限会社"]
                            supplier_name = reduce(lambda prev, current: prev.replace(current, ""), replace_key, supplier_name)
                            # フォントの設定
                            font_name = "BIZ-UDGothicR"
                            pdfmetrics.registerFont(TTFont("BIZ-UDGothicR", r"C:\Windows\Fonts\BIZ-UDGothicR.ttc"))
                            cc.setFont(font_name ,size=10)
                            cc.drawString(target_x, target_y, supplier_name)

                        target_y -= 27 
                        break

           
                print("==========================================================================================")
            
            # ページ挿入
            target_y = 635
            cc.showPage()

        # 保存
        cc.save()
                
                    

        text_file = PdfFileReader(insert_text_file_path, "rb")
        input_file = PdfFileReader(ipo.pdf_files, strict=False)
        page_count = input_file.getNumPages()

        # 新規の出力ファイル作成
        output_file = PdfFileWriter()
        # 既存の全体ページをループで回す
        for page_num in range(page_count):

            input_page = input_file.getPage(page_num)
            # 既存のページとテキストをmergeする
            input_page.mergePage(text_file.getPage(page_num))
            # 圧縮する
            input_page.compressContentStreams()
            # 出力ファイルにページを追加する
            output_file.addPage(input_page)
        # 出力ファイル保存
        with open(f"{insert_text_file_path[:-6]}_new.pdf", "wb") as outputStream:
            output_file.write(outputStream)
        

        return success_action()
    except Exception as e:
        return error_action(e)
   

  
     
if __name__ == "__main__":
    print(123)
   