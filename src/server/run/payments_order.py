from dataclasses import dataclass, field
from datetime import datetime, timedelta

from dotenv import load_dotenv
from server.function import datetime_to_str, get_various_days, list_to_merge_str, today
from typing import Any, Literal, TypedDict
import os
import signal
from openpyxl import Workbook
from openpyxl import styles
from collections import Counter
from server.classes.Obi import Obi
from server.function import error_action, get_various_consecutive_date_list, str_to_datetime, success_action, today_str
from server.type import ExcelDataKoutei, KouteiDataType
from server.utils import get_id 
from server.classes.Op import Op 
from server.classes.OpKoutei import OpKoutei
from server.classes.OpPayments import OpPayments
from server.classes.Rd import Rd

# load_dotenv("../../.env")

class PaymentsItemInfo(TypedDict):
    is_need_order: bool 
    parent_item_num: str
    parent_delivery_time: str
    payments_item_num: str
    payments_item_name:str 
    payments_item_spec: str
    order_num: str
    lines_num: str
    base_qty: int
    qty: int

class DuplicateItemInfo(TypedDict):
    item_num: str
    qty: int
    is_need_order: bool

class PoExcelListType(TypedDict, total=False):
    受注NO: str
    行NO: str
    親品番: str
    親品名: str
    親数量: int
    親取引先code: str
    親取引先名: str
    親納期: str
    親工程数: int
    支給品番: str
    支給品名: str
    支給取引先名: str
    支給元数量: int
    支給手配数量:int
    未引当数: int
    実利1: int
    支給納期: str
    EDI: Literal["はい", "いいえ"]
    手配有無: Literal["有り", "無し"]
    内示: Literal["済", "未"]

class OpDataType(TypedDict):
    order_num: str
    lines_num: str
    item_num: str
    item_name: str
    qty: int

class PaymentsKouteiDataType(TypedDict):
    取引先code: str
    取引先名: str
    納期: str
    工程数: int
    

@dataclass
class PaymentsOrder(Op):
    id: str
    password: str
    startPage: int
    type: Literal["国内", "海外"] = "国内"
    nameInitial: str = ""
    
    op_data_lib: Any = field(default_factory=dict)
    koutei_data_lib: Any = field(default_factory=dict)
    excel_data_list: list[PoExcelListType] = field(default_factory=list)    
    excel_lib: Any = field(default_factory=dict)
    payments_item_list: list[PaymentsItemInfo] = field(default_factory=list)    
    duplicate_item_list: list[DuplicateItemInfo] = field(default_factory=list)    

    def __post_init__(self) -> None:
        super().set_up()
        super().open_main_page(os.environ["SERVICE_SYSTEM_URL"])
        super().Login(self.id, self.password)
    
    def get_op_data(self, i: int) -> None:
        self.op_data_lib = {}
        op_data: OpDataType = {
            "order_num": super().get_value(get_id("受注NO_op_results", i)),
            "lines_num": super().get_value(get_id("行NO_op_results", i)),
            "item_num": super().get_value(get_id("品番_op_results", i)),
            "item_name": super().get_value(get_id("品名_op_results", i)),
            "qty": int(super().get_value(get_id("数量_op_results", i)).replace(',', ''))
        }

        self.op_data_lib = op_data
    
    def get_koutei_data(self, koutei_excel_data: ExcelDataKoutei) -> None:
        self.koutei_data_lib = {}
        koutei_data: PaymentsKouteiDataType = {
            "取引先code": koutei_excel_data["取引先code"],
            "取引先名": koutei_excel_data["取引先名"],
            "納期": koutei_excel_data["納期"],
            "工程数": koutei_excel_data["工程数"],
        }

        self.koutei_data_lib = koutei_data

    def excel_data_collecting_op(self) -> None:
        excel_data: PoExcelListType = {}
        excel_data["受注NO"] = self.op_data_lib["order_num"]
        excel_data["行NO"] = self.op_data_lib["lines_num"]
        excel_data["親品番"] = self.op_data_lib["item_num"]
        excel_data["親品名"] = self.op_data_lib["item_name"]
        excel_data["親数量"] = self.op_data_lib["qty"]
        
        self.update_excel_lib(excel_data)
    
    
    def excel_data_collecting_koutei(self) -> None:
        excel_data: PoExcelListType = {}
        excel_data["親取引先code"] = self.koutei_data_lib["取引先code"] 
        excel_data["親取引先名"] = self.koutei_data_lib["取引先名"]  
        excel_data["親納期"] = self.koutei_data_lib["納期"] 
        excel_data["親工程数"] = self.koutei_data_lib["工程数"]
        
        self.update_excel_lib(excel_data)
    
    def excel_data_collecting_payments(self, item_num: str, item_name: str, qty: int) -> None:
        excel_data: PoExcelListType = {}
        excel_data["支給品番"] = item_num
        excel_data["支給品名"] = item_name
        excel_data["支給元数量"] = qty

        self.update_excel_lib(excel_data)

    def excel_data_collecting_obi(self, payments_data: PaymentsItemInfo, total_qty: int) -> None :
        excel_data: PoExcelListType = {}
        excel_data["手配有無"] = "有り" if payments_data["is_need_order"] else "無し"
        excel_data["支給手配数量"] = payments_data["qty"]
        excel_data["未引当数"] = total_qty
        excel_data["実利1"] = int(super().get_value(get_id("実利1_obi_results")).replace(',', ''))

        self.update_excel_lib(excel_data)
    

    def update_excel_lib(self, excel_data: PoExcelListType) -> None:
        self.excel_lib.update(excel_data)
    
    def append_excel_list(self)-> None:
        self.excel_data_list.append(self.excel_lib)
        self.excel_lib = {}

    def append_duplicate_list(self, data: DuplicateItemInfo)-> None:
        self.duplicate_item_list.append(data)
    
    def append_payments_list(self, data: PaymentsItemInfo)-> None:
        self.payments_item_list.append(data)
    

    

    def new_excel(self):
        downloads_path = os.getenv("HOMEDRIVE") + os.getenv("HOMEPATH") + "\\Downloads"
        header_list = ["受注NO","行NO","親品番","親品名","親数量","親取引先code", "親取引先名", "親納期", "親工程数","支給品番","支給品名","支給取引先名", "EDI","支給元数量", "支給手配数量", "未引当数", "実利1","支給納期", "手配有無", "内示"]
        wb = Workbook()
        ws = wb.active
        # header作成
        for i, title in enumerate(header_list):
            ws.cell(1,i+1).value = title
        
        for y, data in enumerate(self.excel_data_list):
            for x, title in enumerate(header_list):
                try:
                    ws.cell(y+2, x+1).value = data[title]
                except KeyError:
                    ws.cell(y+2, x+1).value = ""
        # ﾌｨﾙﾀｰ設定
        ws.auto_filter.ref = f"A1:X{ws.max_row}"
        # header固定
        ws.freeze_panes = 'A2'


        fill1 = styles.PatternFill(patternType='solid',fgColor='C0C0C0', bgColor='C0C0C0')
        fill2 = styles.PatternFill(patternType='solid',fgColor='FFFF00', bgColor='FFFF00')
        # header色変更
        rows = ws["A1":"X1"]
        for row in rows[0]:
            row.fill = fill1

        Item_num_list = [ws.cell(i+2, 10).value for i in range(ws.max_row-1)]
        duplicate_item_list = [key for key,value in Counter(Item_num_list).items() if value > 1]

        for i in range(ws.max_row-1):
            current_cell = ws.cell(i+2, 10)
            if (current_cell.value in duplicate_item_list):
                current_cell.fill = fill2
        
        wb.save(f"{downloads_path}\支給品{today_str(add_time=True)}.xlsx")
    

    def get_payments_item_data(self, payments: OpPayments, is_need_order: bool, payments_item_num: str, payments_item_name: str, payments_item_spec: str,  qty: int) -> PaymentsItemInfo:
        return {
            "is_need_order": is_need_order, 
            "parent_item_num": payments.item_num,
            "parent_delivery_time":  self.koutei_data_lib.get("納期") if self.koutei_data_lib.get("納期") else "",
            "payments_item_num": payments_item_num, 
            "payments_item_name": payments_item_name,
            "payments_item_spec": payments_item_spec,  
            "order_num": payments.order_num,
            "lines_num": payments.lines_num,
            "base_qty": self.excel_lib.get("支給元数量") if self.excel_lib.get("支給元数量") else 0 ,
            "qty": qty
        }

    def parent_KOUTEI_process(self, i: int) -> None:
        koutei = OpKoutei(self.brower, self.wait, i, self.type)
        koutei.start()
        koutei_excel_data = koutei.get_data_for_excel()
        koutei.end()

        self.get_koutei_data(koutei_excel_data)

    def payments_KOUTEI_process(self, i: int, parent_delivery_time: str) -> list[KouteiDataType]:
        
        koutei = OpKoutei(self.brower, self.wait, i, self.type)
        koutei.start()
        koutei.set_price_not_setting()

        delivery_time_list = self.decide_payments_delivery_time(koutei.num, parent_delivery_time)

        for i in range(koutei.num):
            super().set_value(get_id("納期_op_KOUTEI", i), delivery_time_list[i])

        koutei_data = koutei.get_data() 
        koutei.end()

        return koutei_data

    
    def decide_payments_delivery_time(self, koutei_num: int, parent_delivery_time: str) ->list[str]:
        split_delivery_time_str = parent_delivery_time.split("__")
        delivery_time_date: datetime = str_to_datetime(split_delivery_time_str[0])
        later_weeks_list = get_various_consecutive_date_list("move", 6)
        four_days_later = get_various_days("move", 4)
        
        if koutei_num == 1:
            for i in range(4): 
                if later_weeks_list[3 - i] < delivery_time_date:
                    delivery_time_date_list = [later_weeks_list[2 - i]] if not i == 3 else [four_days_later]
                    break
            else:
                delivery_time_date_list = [four_days_later]

        elif koutei_num == 2:
            if later_weeks_list[5] < delivery_time_date:
                delivery_time_date_list = [later_weeks_list[1] + timedelta(weeks=i) for i in range(3)]
            
            else:
                delivery_time_date_list = [later_weeks_list[0] + timedelta(weeks=i) for i in range(2)]

        else:
            delivery_time_date_list = [later_weeks_list[0] + timedelta(weeks=i) for i in range(3)]

        
        delivery_time_str_list: list[str] = [datetime_to_str(delivery_time) for delivery_time in delivery_time_date_list]

        
        return delivery_time_str_list

    
    def payments_process(self, i: int, order_num: str, lines_num: str, item_num: str) -> None:
        payments = OpPayments(self.brower, self.wait, i, order_num, lines_num, item_num)
        payments.start()

        if super().check_pages_exists():
            is_not_last_page = True
            while is_not_last_page:
                
                self.payments_main_process(payments)
                        
                is_not_last_page = super().check_next_page_exists()
                if is_not_last_page:
                    super().move_next_page()
                    super().the_element_view_wait(get_id("品番_op_payments", 0))

        
        # 支給メニューにﾍﾟｰｼﾞが無い場合
        else:
            self.payments_main_process(payments)
                

        payments.end()

    
    def payments_main_process(self, payments: OpPayments):
        payments_item_num = sum(payments.payments_exists()) 
        for i in range(payments_item_num):

            payments_item_num = super().get_value(get_id("品番_op_payments", i))
            payments_item_name = super().get_value(get_id("品名_op_payments", i)) 
            payments_item_spec = super().get_value(get_id("仕様_op_payments", i)) 
            current_qty = int(super().get_value(get_id("数量_op_payments", i)).replace(",", ""))

            self.excel_data_collecting_payments(payments_item_num, payments_item_name, current_qty)

            # 同品番での手配が既にあるかをチェックし、合計数量を計算--------------------            
            duplicate_items = [item for item in self.duplicate_item_list if item["item_num"] == payments_item_num]
            duplicate_qty_list = [item["qty"] for item in duplicate_items]
            duplicate_total_qty = sum(duplicate_qty_list)
            # ------------------------------------------------------------------------

            # 品目別受注照会へ----------------------------------------------------------
            obi = Obi(self.brower, self.wait, "", "", payments_item_num, 2)
            is_result = obi.search()
            ZITURI_1 = int(super().get_value(get_id("実利1_obi_results")).replace(",", ""))

            if not is_result:
                is_need_order = current_qty + duplicate_total_qty > ZITURI_1
                self.update_all_list(payments, is_need_order, payments_item_num, payments_item_name, payments_item_spec, current_qty, duplicate_total_qty)
                super().screen_switching(1)
                continue
            # -------------------------------------------------------------------------


            # 検索結果に地方拠点の補充が含まれている場合は打切----------------------------
            is_check_replenishment = obi.check_replenishment()
            if is_check_replenishment:
                rd = Rd(self.brower, self.wait, "", "", payments_item_num, 3)
                rd.nomal_process()
                super().screen_switching(2)
                is_result = obi.search()
                if not is_result:
                    is_need_order = current_qty + duplicate_total_qty > ZITURI_1
                    self.update_all_list(payments, is_need_order, payments_item_num, payments_item_name, payments_item_spec, current_qty, duplicate_total_qty)
                    super().screen_switching(1)
                    continue
            # -------------------------------------------------------------------------

            # 品目別受注照会の検索結果がある場合
            obi_result = obi.check_all_order_situation(return_qty=True)
            obi_total_qty = obi_result["data"]["total_qty"]
           
            # 既に同じ部品を手配しているかどうか     
            total_qty =  duplicate_total_qty if duplicate_total_qty else  obi_total_qty      
            qty = current_qty + total_qty
            is_need_order = qty > ZITURI_1
            self.update_all_list(payments, is_need_order, payments_item_num, payments_item_name, payments_item_spec, qty, total_qty)

            
            super().screen_switching(1)

    

    def update_all_list(self, payments: OpPayments, is_need_order: bool, payments_item_num: str, payments_item_name: str, payments_item_spec: str,  qty: int, total_qty: int) -> None:
        payments_data = self.get_payments_item_data(payments, is_need_order, payments_item_num, payments_item_name, payments_item_spec, qty)
        duplicate_item: DuplicateItemInfo = {
            "item_num": payments_item_num,
            "qty": payments_data["qty"],
            "is_need_order": payments_data["is_need_order"]
        }
        self.excel_data_collecting_op()
        self.excel_data_collecting_koutei()
        self.excel_data_collecting_obi(payments_data, total_qty)
        self.append_payments_list(payments_data)
        self.append_excel_list()
        self.append_duplicate_list(duplicate_item)
    

    def can_items_order(self, order_num: str, lines_num: str, item_num: str) -> bool:    
        current_payments_items: list[PaymentsItemInfo] = list(filter(
            lambda item: 
                item["parent_item_num"] == item_num 
                and item["order_num"] == order_num 
                and item["lines_num"] ==lines_num 
            , self.payments_item_list
        ))

        is_need_order_items = [item for item in current_payments_items if item["is_need_order"] ]
        can_order = False if len(is_need_order_items) else True

        return can_order
    

    def insert_payments_koutei_data(self, payments_item: PaymentsItemInfo, koutei_data: list[KouteiDataType], can_payments_item_orders: bool) -> None:
        def map_func(excel_data: PoExcelListType) -> Any:
            if (excel_data["受注NO"] == payments_item["order_num"] 
                and excel_data["行NO"] == payments_item["lines_num"] 
                and excel_data["支給品番"] == payments_item["payments_item_num"]
                and excel_data["支給手配数量"] == payments_item["qty"]):

                supplier_name = list_to_merge_str(koutei_data, "/", "取引先名") 
                delivery_time = list_to_merge_str(koutei_data, "/", "納期") 

                EDI_list = [koutei["EDI"] for koutei in koutei_data]
                EDI_text = "いいえ" if False in EDI_list else "はい"

                NAIJI_text = "済" if can_payments_item_orders else "未"

                inserted_excel_data: Any  = {
                    **excel_data, 
                    "支給取引先名": supplier_name, 
                    "支給納期": delivery_time,
                    "EDI": EDI_text,
                    "内示": NAIJI_text
                    }
                
                
                return inserted_excel_data

            return excel_data

        inserted_excel_data_list = list(map(map_func, self.excel_data_list))
        self.excel_data_list = inserted_excel_data_list
    

    def decide_payments_comment(self, koutei_data: list[KouteiDataType], payments_item: PaymentsItemInfo , is_payments: bool) -> str:
        EDI_list = [koutei["EDI"] for koutei in koutei_data]
        is_EDI = all(EDI_list)
        beginning_text = "紙 支給 " if is_payments and not is_EDI  else "支給 " if is_payments else  "紙 " if not is_EDI else  f"{today_str()} "
        base_text = f"{payments_item['parent_item_num']}への支給品  {self.nameInitial}"

        KOUTEI_info = " ".join([f"工程({i+1}):{data['取引先名']}" for i, data in enumerate(koutei_data)])
        KOUTEI_text = f"{KOUTEI_info} " if len(koutei_data) > 1 else " "

        comment = f"{beginning_text}{KOUTEI_text}{base_text}"

        return comment







# =====↓↓処理↓↓==============================================================================

def payments_order(id: str, password: str,  type: Literal["国内", "海外"] = "国内", startPage: int = 0, nameInitial: str = ""):
    po = PaymentsOrder(id, password, startPage, type, nameInitial)
    try:
        po.menu_open("op_entry", 1)
        po.menu_open("obi_entry", 2)
        po.menu_open("rd_entry", 3)
        po.screen_switching(1)
        po.header_input(po.type)
        po.move_start_page(startPage)


        is_not_last_page = True
        while is_not_last_page:

            items_list = po.get_item_els()
            for i, item in enumerate(items_list):
                if item == []:
                    break

                #備考欄に”支給”がある場合
                comment = po.get_value(get_id("備考_op_results", i))
                if "支給" in  comment:
                    current_order_num = po.get_value(get_id("受注NO_op_results", i))
                    current_lines_num = po.get_value(get_id("行NO_op_results", i))
                    parent_item_num = po.get_value(get_id("品番_op_results", i))
 
                    po.get_op_data(i)
                    po.parent_KOUTEI_process(i)
                    po.payments_process(i, current_order_num, current_lines_num, parent_item_num)

                    remove_payments_text = comment.replace("支給", "")
                    if po.can_items_order(current_order_num, current_lines_num, parent_item_num):
                        if "紙" not in comment:
                            new_comment = f"{today_str()} 支給品在庫有り {remove_payments_text}"
     
                            po.set_value(get_id("備考_op_results", i), new_comment)
                            po.btn_click(get_id("更新_op_results", i))
                            po.btn_click(get_id("内示_op_results", i))
                    else:
                        if "紙" not in comment:
                            new_comment = f"{today_str()} 支給品納期確認中 {remove_payments_text}"
                            po.set_value(get_id("備考_op_results", i), new_comment)
                            po.btn_click(get_id("更新_op_results", i))

                 
            is_not_last_page = po.check_next_page_exists()

            if is_not_last_page:
                po.move_next_page()
                po.nomal_wait()
        

        
        is_completed = po.complete()
        if not is_completed:
            po.btn_click(get_id( "クリア_op_search"))
            po.alert_accept()
            po.clickable_wait("H_srch_btn")
            print("親確定時にエラー発生")
        # ＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝
        


        # ＝＝＝支給品の発注計画新規作成＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝

        po.btn_click(get_id("新規追加_radio_btn_op_search"))
        po.btn_click(get_id("新規追加_op_search"))
        po.the_element_view_wait("M1")
        
        
        is_need_orders: list[bool] = [ item["is_need_order"] for item in po.payments_item_list]
        if any(is_need_orders):
            idx = 0 
            for payments_item in po.payments_item_list:
                if payments_item["is_need_order"]:
                    po.btn_click(get_id("更新_op_results", idx))
                    po.set_value(get_id("品番_op_results", idx), payments_item["payments_item_num"])
                    po.set_value(get_id("受注NO_op_results", idx), payments_item["order_num"])
                    po.set_value(get_id("数量_op_results", idx), payments_item["qty"])
                    po.set_value(get_id("製番_op_results", idx), "9300NAI")


                    koutei_data = po.payments_KOUTEI_process(idx, payments_item["parent_delivery_time"])

                    # po.insert_payments_koutei_data(payments_item, koutei_data)

                    payments = OpPayments(po.brower, po.wait, idx)
                    is_payments = payments.click_only_when_creating_op()

                    comment = po.decide_payments_comment(koutei_data, payments_item, is_payments)
                    po.set_value(get_id("備考_op_results", idx), comment)

                    EDI_list = [koutei["EDI"] for koutei in koutei_data]
                    is_EDI = all(EDI_list)
                    is_same_qty = payments_item["base_qty"] == payments_item["qty"]

                    alternative_item_exists = any([keyword in payments_item["payments_item_spec"] for keyword in ["*1*", "*2*"]])

                    can_payments_item_orders = not is_payments and is_EDI and is_same_qty and not alternative_item_exists

                    if can_payments_item_orders:

                        po.btn_click(get_id("内示_op_results", idx))

                    po.insert_payments_koutei_data(payments_item, koutei_data, can_payments_item_orders)

                    idx += 1

                    if idx > 3:
                        idx = 0
                    
                    po.btn_click(get_id("明細追加_op_footer"))
                    po.the_element_view_wait(get_id("品番_op_results", idx))
                
            po.btn_click(get_id("再表示_op_footer"))
            po.clickable_wait("H_cnfrm_btn")
            po.btn_click(get_id("確認_op_footer"))
            po.clickable_wait("H_dtrmn_btn")
            po.btn_click(get_id("確定_op_footer"))
       
       # ＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝

        po.new_excel()

# ========================================================================================
        return success_action()
    except Exception as e:
        po.new_excel()
        return error_action(e)
    finally:
        os.kill(po.brower.service.process.pid,signal.SIGTERM)

  
     
if __name__ == "__main__":
    print(123)
