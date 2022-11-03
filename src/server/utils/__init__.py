from typing import Any, Final, Literal, Optional
from openpyxl import Workbook
from openpyxl import styles

def get_id(key: Any, i: int=0) ->  Optional[str]:
    id_text: Final[str] = f"{i+2}" if len(str(i+2)) == 2 else f"0{i+2}"

    element_ids: Final[dict[str, str]] = {
        "新規追加_radio_btn_op_search": "H_dspsl_mode_1",
        "受注NO_op_search": "H_ordr_slip_no",
        "品番_op_search": "H_item_cd",
        "計画区分_op_search": "H_plan_kbn",
        "内示_op_search": "H_insd_flg",
        "検索_op_search": "H_srch_btn",
        "クリア_op_search": "H_clear_btn",
        "新規追加_op_search": "H_new_btn",

        "受注NO_op_results": f"M1_ctl0{i+2}_H_m1_ordr_slip_no",
        "行NO_op_results": f"M1_ctl0{i+2}_H_m1_ordr_slip_dtail_no",
        "品番_op_results": f"M1_ctl0{i+2}_H_m1_item_cd",
        "品名_op_results": f"M1_ctl0{i+2}_H_m1_item_nm",
        "仕様_op_results": f"M1_ctl0{i+2}_H_m1_spec",
        "数量_op_results": f"M1_ctl0{i+2}_H_m1_po_qty",
        "LOT_op_results": f"M1_ctl0{i+2}_H_m1_purch_lot",
        "科目_op_results": f"M1_ctl0{i+2}_H_m1_titl_cd",
        "L/T_op_results": f"M1_ctl0{i+2}_H_m1_item_parts_lt",
        "製番_op_results": f"M1_ctl0{i+2}_H_m1_prdt_no",
        "備考_op_results": f"M1_ctl0{i+2}_H_m1_note",
        "緊急度_op_results": f"M1_ctl0{i+2}_H_m1_urgnt_degree_kbn",
        "支給品ﾌﾗｸﾞ_op_results": f"M1_ctl0{i+2}_H_m1_prvd_kbn",
        "生産中止ﾌﾗｸﾞ_op_results": f"M1_ctl0{i+2}_H_m1_mnfct_end_flg",
        "EDIﾌﾗｸﾞ_op_results": f"M1_ctl0{i+2}_H_m1_edi_trad_kbn",
        "受注者ｺｰﾄﾞ_op_results": f"M1_ctl0{i+2}_H_m1_ordr_prsn_cd",
        "工程_op_results": f"M1_ctl0{i+2}_H_m1_wp_btn",      
        "支給_op_results": f"M1_ctl0{i+2}_H_m1_prvd_btn",      
        "更新_op_results": f"M1_ctl0{i+2}_H_m1_upd_chk",
        "削除_op_results": f"M1_ctl0{i+2}_H_m1_clsr_chk",
        "内示_op_results": f"M1_ctl0{i+2}_H_m1_insd_chk",
        "確定check_op_results": f"M1_ctl0{i+2}_H_m1_fix_chk",
        "確認_op_results": f"H_cnfrm_btn",
        "確定_op_results": f"H_dtrmn_btn",
        
        "工程NO_op_KOUTEI" : f"M1_ctl0{i+2}_H_m1_wrk_prcs_no",
        "単価_op_KOUTEI" : f"M1_ctl0{i+2}_H_m1_po_untprc",
        "単区_op_KOUTEI" : f"M1_ctl0{i+2}_H_m1_untprc_kbn",
        "単価未設定_op_KOUTEI" : f"M1_ctl0{i+2}_H_m1_untprc_noset_rsn",
        "納期_op_KOUTEI" : f"M1_ctl0{i+2}_H_m1_dlvry_days",
        "確認_op_KOUTEI" : f"H_cnfrm_btn",
        "確定_op_KOUTEI" : f"H_dtrmn_btn",
        "取引先code_op_KOUTEI" : f"M1_ctl0{i+2}_H_m1_wkcnt_cd",
        "取引先名_op_KOUTEI": f"M1_ctl0{i+2}_H_m1_wkcnt_nm",
        "EDI_op_KOUTEI": f"M1_ctl0{i+2}_H_m1_edi_trad_kbn",
        
        "品番_op_payments" : f"M1_ctl0{i+2}_H_m1_item_cd",
        "品名_op_payments" : f"M1_ctl0{i+2}_H_m1_item_nm",
        "仕様_op_payments" : f"M1_ctl0{i+2}_H_m1_spec",
        "数量_op_payments" : f"M1_ctl0{i+2}_H_m1_out_wrhos_rqst_qty",
        "確認_op_payments" : "H_cnfrm_btn",
        "確定_op_payments" : "H_dtrmn_btn",

        "拠点_obi_search": "H_base_point_cd",
        "品番_obi_search": "H_item_cd",
        "品名_obi_search": 'H_item_nm',
        "仕様_obi_search": "H_spec",
        "検索_obi_search": "H_srch_btn",

        "確認_op_footer" : "H_cnfrm_btn",
        "確定_op_footer" : "H_dtrmn_btn",
        "明細追加_op_footer" : "H_dtl_add_btn",
        "再表示_op_footer" : "H_reload_btn",
        
        "受注NO_obi_results": f'M1_ctl0{i+2}_H_m1_ordr_slip_no',
        "行NO_obi_results": f'M1_ctl0{i+2}_H_m1_ordr_ln_no',
        "数量_obi_results": f"M1_ctl0{i+2}_H_m1_qty",
        "製番_obi_results": f'M1_ctl0{i+2}_H_m1_prdt_no',
        "受注担当者_obi_results": f'M1_ctl0{i+2}_H_m1_ordr_prsn_nm',
        "ｽﾃｰﾀｽ_obi_results": f'M1_ctl0{i+2}_H_m1_stts',
        "受注数_obi_results": f'M1_ctl0{i+2}_H_m1_qty',
        "発注数_obi_results": f'H_po_qty',
        "実利1_obi_results": f'H_actual_useful_qty1',
        "実利2_obi_results": f'H_actual_useful_qty2',
        "区分_obi_results": f'M1_ctl0{i+2}_H_m1_kbn',
        "修理日_obi_results": f'M1_ctl0{i+2}_H_m1_rpr_schdl_dt',
        "客先ｺｰﾄﾞ_obi_results": f'M1_ctl0{i+2}_H_m1_dlvry_to_cd',
        "客先名_obi_results": f'M1_ctl0{i+2}_H_m1_dlvry_to_nm',
        "取引種別_obi_results": f'M1_ctl{id_text}_H_m1_dlng_type',

        "品番_ha_search": "H_item_cd",
        "検索_ha_search": "H_srch_btn",

        "受注NO_ha_results": f'M1_ctl{id_text}_H_m1_ordr_slip_no',
        "行NO_ha_results": f'M1_ctl{id_text}_H_m1_ordr_ln_no',
        "変更_ha_results": f'M1_ctl{id_text}_H_m1_rfrm_btn',
        "削除_ha_results": f'M1_ctl02_H_m1_slct_chk',
        "確認_ha_results" : "H_cnfrm_btn",
        "確定_ha_results" : "H_dtrmn_btn",

        "品番_rd_header": "header_Button2",

        "品番_rd_search": "H_item_cd",
        "検索_rd_search": "H_srch_btn",
        "確認_rd_results" : "H_cnfrm_btn",
        "確定_rd_results" : "H_dtrmn_btn",

        "削除_rd_results": f"M1_ctl{id_text}_H_m1_slct_chk",
        "ｽﾃｰﾀｽ_rd_results": f"M1_ctl{id_text}_H_m1_ship_stts",

        "確認_rd_footer" : "H_cnfrm_btn",
        "確定_rd_footer" : "H_dtrmn_btn",

        "検索_ij_search" : "H_srch_btn",

        "受注NO_ij_results": f'M1_ctl0{i+2}_H_m1_ordr_slip_no',
        "行NO_ij_results": f'M1_ctl0{i+2}_H_m1_ordr_slip_dtail_no',
        "品番_ij_results" : f"M1_ctl0{i+2}_H_m1_item_cd",
        "品名_ij_results" : f"M1_ctl0{i+2}_H_m1_item_nm",
        "仕様_ij_results" : f"M1_ctl0{i+2}_H_m1_spec",
        "受注者_ij_results" : f"M1_ctl0{i+2}_H_m1_ordr_prsn_nm",
        "ｺﾒﾝﾄ_ij_results" : f"M1_ctl0{i+2}_H_m1_dtmn_comnt",
        "DBﾌﾗｸﾞ_ij_results" : f"M1_ctl0{i+2}_H_m1_db_kbn",
        "更新_ij_results" : f"M1_ctl0{i+2}_H_m1_upd_chk",
        "確定_ij_results" : f"M1_ctl0{i+2}_H_m1_fix_chk",


    }
    
    element_id = element_ids.get(key)
    if element_id:
        return element_id
    else:
        return None

def get_xpath(key: Any) -> Optional[str]:
    element_ids = {
        "ﾍﾟｰｼﾞ_op_results": "//*[@id='H_pager']/b",
        "ﾍﾟｰｼﾞ_Next_btn_op_results": "//*[@id='H_pager']/a",

        "検索結果_obi_results": "//*[@id='M1']/tbody/tr",
        "ﾍﾟｰｼﾞ_Next_btn_obi_results": "//*[@id='H_pager']/a",

        "検索結果_ha_results": "//*[@id='M1']/tbody/tr",

        "検索結果_rd_results": "//*[@id='M1']/tbody/tr",

        "ﾍﾟｰｼﾞ_Next_btn_ij_results": "//*[@id='H_pager']/a",

        "ﾍﾟｰｼﾞ_Next_btn": "//*[@id='H_pager']/a",
    }

    xpath_text = element_ids.get(key)
    if xpath_text:
        return xpath_text
    else:
        return None



def create_excel(excel_data_list: list, header_list: list[str], file_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    # header作成
    for i, title in enumerate(header_list):
        ws.cell(1,i+1).value = title
    
    for y, data in enumerate(excel_data_list):
        for x, title in enumerate(header_list):
            try:
                ws.cell(y+2, x+1).value = data[title]
            except KeyError:
                ws.cell(y+2, x+1).value = ""
    # ﾌｨﾙﾀｰ設定
    ws.auto_filter.ref = f"A1:AI{ws.max_row}"
    # header固定
    ws.freeze_panes = 'A2'

    fill1 = styles.PatternFill(patternType='solid',fgColor='C0C0C0', bgColor='C0C0C0')

    # header色変更
    for i in range(len(header_list)):
        cell = ws.cell(1, i+1)
        cell.fill = fill1
    
    wb.save(file_path)