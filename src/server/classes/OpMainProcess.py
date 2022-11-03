import os
from server.classes.Op import Op
from server.classes.OpKoutei import OpKoutei
from server.classes.OpPayments import OpPayments
from server.classes.Obi import Obi
from server.classes.Rd import Rd
from typing import Final, Literal, Optional, TypeGuard, Any
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl import styles
import collections
from server.type import ExcelDataObi, ObiOrderExistsResultsType, ObiOrderExistsType, RuleItemsType
from server.type import ExcelDataKoutei, ExcelDataOp, ObiOrderSituationResultsType, ObiOrderSituationType
from server.function import str_to_datetime, get_download_path, today_str
from server.utils import get_id


@dataclass
class OpMainProcess(Op):
    id: str 
    password: str
    startPage: int = 0
    nameInitial: str = ""
    settings_items: list[RuleItemsType] = field(default_factory=list)
    op_type: Literal["国内", "海外"] = "国内"

    excel_list: Any = field(default_factory=list)
    excel_lib: Any = field(default_factory=dict)
    duplicate_item_list: Any  = field(default_factory=list)
    duplicate_item_lib: Any = field(default_factory=dict)

    def __post_init__(self) -> None:
        super().set_up()
        super().open_main_page(os.environ["SERVICE_SYSTEM_URL"])
        super().Login(self.id, self.password)

    def confirmed_as_it_is(self, i: int, notes_text: str="", is_excel: bool=True) -> None:  
        super().btn_click(get_id("更新_op_results", i))
        super().set_value(get_id("備考_op_results", i), notes_text)
        super().btn_click(get_id("確定check_op_results", i))
        koutei = OpKoutei(self.brower, self.wait, i)
        koutei.click_only()
        payments = OpPayments(self.brower, self.wait, i)
        payments.click_only()
        
        if is_excel:
            excel_data = super().get_data_for_excel(i, "確定")
            self.update_excel_lib(excel_data)
            self.add_to_excel_list()
        

    def update_excel_lib(self, excel_data: ExcelDataOp | ExcelDataObi | ExcelDataKoutei) -> None:
        self.excel_lib.update(excel_data)
    
    def add_to_excel_list(self) -> None:
        self.excel_list.append(self.excel_lib)
        self.excel_lib = {}
    
    def update_duplicate_item_lib(self, item_data) -> None:
        self.duplicate_item_lib.update(item_data)
    
    def add_to_duplicate_item_list(self) -> None:
        self.duplicate_item_list.append(self.duplicate_item_lib)
        self.duplicate_item_lib = {}
    
    def delete_and_add_excel_data(self, i: int, notes_text: str="") -> None:
        super().delete(i, notes_text)
        excel_data = super().get_data_for_excel(i, "削除")
        self.update_excel_lib(excel_data)
        self.add_to_excel_list()
    
    def is_ObiOrderSituationResultsType(self, result_obi: ObiOrderSituationType | ObiOrderExistsType) -> TypeGuard[ObiOrderSituationType]:
        if "many" in result_obi:
            return True
       
        return False
    

    def process(self, i: int, process_type: Literal["通常", "引当"]="通常") -> None:
        current_op_data = super().get_required_data_lib(i)

        # 品目別受注照会へ
        obi = Obi(self.brower, self.wait, current_op_data["受注NO"], current_op_data["行NO"], current_op_data["品番"], 2)
        is_result = obi.search()

        if not is_result:
            super().screen_switching(1)
            self.delete_and_add_excel_data(i, "受注残無し")  
            
            return


        if process_type == "通常":
            is_check_replenishment = obi.check_replenishment()

            if is_check_replenishment:
                rd =Rd(self.brower, self.wait, obi.op_order_num, obi.op_lines_num, obi.op_item_num, 4)
                rd.nomal_process()
                super().screen_switching(2)
                is_result = obi.search()
                if not is_result:
                    super().screen_switching(1)
                    self.delete_and_add_excel_data(i, "web 受注残無し")  
                    
                    return
         
        result: ObiOrderSituationResultsType | ObiOrderExistsResultsType   = (
             obi.check_order_situation()  if process_type == "通常" else obi.check_all_order()
        )
        
        obi_result_data = result["data"]
        obi_excel_data = result["excel"]

        if obi_excel_data:
            self.update_excel_lib(obi_excel_data)

        super().screen_switching(1)

        if process_type == "通常" and self.is_ObiOrderSituationResultsType(obi_result_data):
            # 受注残が多い場合
            if obi_result_data["many"]:  
                super().add_comment_and_through(i, "通常　受注残が多いです")
                return 

            # 引当調整を行った場合
            if obi_result_data["adjust"]:
                self.delete_and_add_excel_data(i, "引当調整対応")

                return 
        
        # 既に受注残が無い場合
        if not obi_result_data["exist"]: 
            self.delete_and_add_excel_data(i, "受注残無し")   

            return 

        # 納期を決める
        delivery_time: str = super().decide_delivery_time(obi_result_data["repair_days"], current_op_data["LT"], self.op_type)      


        if process_type == "通常" and self.is_ObiOrderSituationResultsType(obi_result_data):
            # 同品番での手配が既にあるかをチェック
            item_exists_list = super().check_duplicate_item_exists(current_op_data["品番"], self.duplicate_item_list)
            
            # 数量を決める
            qty = super().decide_qty(i, current_op_data["数量"], obi_result_data["total_qty"], item_exists_list, delivery_time)
            super().set_qty(i, qty)

            # 製番設定(93～の場合は　～NAI)
            if super().get_value(get_id("製番_op_results", i))[:2] == "93":
                super().set_value(get_id("製番_op_results", i), "9300NAI")


        # 更新ボタンをクリック
        super().btn_click(get_id("更新_op_results", i))
        
        # 工程画面へ
        koutei = OpKoutei(self.brower, self.wait, i, self.op_type, delivery_time=delivery_time)
        result_koutei = koutei.process()
        KOUTEI_data = result_koutei["data"]
        KOUTEI_excel_data = result_koutei["excel"]
        self.update_excel_lib(KOUTEI_excel_data)
        
        # 支給画面へ
        payments = OpPayments(self.brower, self.wait, i)
        is_payments = payments.click_only()
        
        # コメントを決める
        comment = super().decide_comment(is_payments, KOUTEI_data, self.nameInitial)
        super().set_value(get_id("備考_op_results", i), comment)
        
        # 紙か支給品有を除き、内示にチェックを入れる
        No_EDI_exists = super().check_No_EDI_exists(KOUTEI_data)
        if not (No_EDI_exists or is_payments):
            super().btn_click(get_id("内示_op_results", i))
            
        
        excel_data = super().get_data_for_excel(i, "内示")
        self.update_excel_lib(excel_data)
        self.add_to_excel_list()


        if process_type == "通常":
            item_num_lib = {
                "品番": current_op_data.get("品番") , 
                "納期": str_to_datetime(delivery_time) 
            }

            self.update_duplicate_item_lib(item_num_lib)
            self.add_to_duplicate_item_list()
    

    def new_excel(self) -> None:
        header_list: Final[list[str]] = [
            "対応","客先名","受注NO","行NO","page","品番","品名",
            "仕様","数量","LOT","未引当数","発注数","受注数","実利2","単価",
            "製番","納期","修理日","LT","区分","工程数","取引先code",
            "取引先名","備考"
        ]
        wb = Workbook()
        ws = wb.active
        # header作成
        for i, title in enumerate(header_list):
            ws.cell(1,i+1).value = title
        
        # データ入力
        for y, data in enumerate(self.excel_list):
            for x, title in enumerate(header_list):
                ws.cell(y+2, x+1).value = data.get(title)
                    
        # ﾌｨﾙﾀｰ設定
        ws.auto_filter.ref = f"A1:X{ws.max_row}"
        # header固定
        ws.freeze_panes = 'A2'

        # 色情報取得
        fill1 = styles.PatternFill(patternType='solid',fgColor='C0C0C0', bgColor='C0C0C0')
        fill2 = styles.PatternFill(patternType='solid',fgColor='FFFF00', bgColor='FFFF00')

        #header色変更
        for i in range(24):
            cell = ws.cell(1, i+1)
            cell.fill = fill1
        
        # 手配品番に被りが有る分に色付け
        Item_num_list = [ws.cell(i+2,6).value for i in range(ws.max_row-1)]
        duplicate_item_list = [key for key,value in collections.Counter(Item_num_list).items() if value > 1]
        for i in range(ws.max_row-1):
            current_cell = ws.cell(i+2,6)
            if current_cell.value in duplicate_item_list:
                current_cell.fill = fill2
        
        # wb.save(f"{get_download_path()}\{today_str(True)}.xlsx")
        wb.save(f"Y:\\530_資材事業課\\パーツセンター\\※GPC_購買部\\発注G\\automation\\result_data\\{self.op_type}発注{today_str(True)}.xlsx")
    
    def get_unique_rule(self, i: int) -> Optional[RuleItemsType]:
        if len(self.settings_items) > 0:
            for rule in self.settings_items:
                if rule["itemNum"] == super().get_value(get_id("品番_op_results", i)) and rule["isApply"]:
                    return rule
        
        return None
    
    def web_order_process(self, i: int) -> None:
        current_op_data = super().get_required_data_lib(i)
        current_lot = int(super().get_value(get_id("LOT_op_results", i)).replace(',', ''))
                
        obi = Obi(self.brower, self.wait, current_op_data["受注NO"], current_op_data["行NO"], current_op_data["品番"], 2)
        is_result = obi.search()
        
        if not is_result:
            super().screen_switching(1)
            self.delete_and_add_excel_data(i, "web 受注残無し")     

            return
        
        is_replenishment = obi.check_replenishment()

        if is_replenishment:
            rd =Rd(self.brower, self.wait, obi.op_order_num, obi.op_lines_num, obi.op_item_num, 4)
            rd.nomal_process()
            super().screen_switching(2)
            is_result = obi.search()
            if not is_result:
                super().screen_switching(1)
                self.delete_and_add_excel_data(i, "web 受注残無し")  
                
                return
        

        obi_result = obi.check_all_order_situation()

        if obi_result["excel"]:
            self.update_excel_lib(obi_result["excel"])

        super().screen_switching(1)

        if not obi_result["data"]:
            super().delete(i, "web 受注残無し")

            return
        
        # 購入ロットが1の場合はそのまま確定
        if current_lot == 1:
            self.confirmed_as_it_is(i, "WEB")

            return
        # 購入ロットが1以上且つ、発注数よりも未引当数が多い場合は適正数量に変更
        if obi_result["data"]["total_qty"] > obi_result["data"]["ordering_qty"]:
            qty = super().check_qty_and_lot(i, current_op_data["数量"])
            super().set_qty(i, qty)
            self.confirmed_as_it_is(i, "WEB")
            
        # 発注数よりも未引当数が少ない場合は、そのまま打ち切りをする
        else:
            self.delete_and_add_excel_data(i, "web 十分足りてる")  



